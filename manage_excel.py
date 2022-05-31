import os
import subprocess
import sys
from os import path
from tkinter import filedialog

from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import json


def open_file(filename):
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, filename])


def verificacion():
    file = filedialog.askdirectory()
    termino = file + '/SubirProductosYapo.xlsx'
    comienzo = False
    i = 1

    while comienzo == False:
        if path.exists(termino) != False:
            termino = file + "/SubirProductosYapo (" + str(i) + ")" + ".xlsx"
            i += 1
        else:
            try:
                comienzo = True
            except:
                break
    return termino


def create_excel_format():
    wb = Workbook()
    hoja = wb.active
    hoja.append(
        (
            'SKU',
            'Categoría (Id)',
            'Título',
            'Descripción',
            'Precio',
            'Imagen 1',
            'Imagen 2',
            'Imagen 3',
            'Imagen 4',
            'Imagen 5',
            'Imagen 6',
            'Región (Id)',
            'Comuna (Id)',
        )
    )

    for i in range(13):
        hoja.column_dimensions[get_column_letter(i + 1)].width = 15

    for rows in hoja.iter_rows(min_row=1, max_row=1):
        for row in rows:
            row.font = Font(bold=True, italic=True)

    ws2 = wb.create_sheet('Región y Comuna')
    ws2.append(('Región', 'Id Región', 'Comuna', 'Id Comuna'))

    for rows in ws2.iter_rows(min_row=1, max_row=1):
        for row in rows:
            row.font = Font(bold=True, italic=True)

    for i in range(4):
        ws2.column_dimensions[get_column_letter(i + 1)].width = 21
    with open('./data/comuna.json', encoding="utf8") as r:
        data = json.loads(r.read())
        for place in data:
            ws2.append((place['region'], place['id_region'], place['comuna'], place['id_comuna']))

    ws3 = wb.create_sheet('Categorías')
    ws3.append(('Id Categoría', 'Categoría 1', 'Categoría 2', 'Categoría 3', 'Categoría 4', 'Categoría 5'))

    for rows in ws3.iter_rows(min_row=1, max_row=1):
        for row in rows:
            row.font = Font(bold=True, italic=True)

    for i in range(6):
        ws3.column_dimensions[get_column_letter(i + 1)].width = 30

    with open('./data/category.json', encoding="utf8") as r:
        data = json.loads(r.read())
        for category in data:
            ws3.append((category, data[category]['categoria_1'], data[category]['categoria_2'],
                        data[category]['condition'], data[category]['gender'], data[category]['clothing_size']))

    filename = verificacion()
    wb.save(filename)
    open_file(filename)


def verification(name):
    file = filedialog.askdirectory()
    termino = file + '/' + name
    comienzo = False
    i = 1

    while comienzo == False:
        if path.exists(termino) != False:
            termino = file + f"/{name} (" + str(i) + ")" + ".xlsx"
            i += 1
        else:
            try:
                comienzo = True
            except:
                break
    return termino


def update_image_excel(file, directory):
    wb = load_workbook(file)
    ws = wb.active
    print(directory)
    for rows in ws.iter_rows(min_row=2, min_col=1):
        code = rows[0].value
        codes = os.listdir(directory)
        if code in codes:
            dir_images = directory + '/' + code
            images = os.listdir(dir_images)
            counter = 1

            if 'default.png' in images:
                rows[5].value = dir_images + '/default.png'
                images.remove('default.png')
                i = 5
            elif 'default.jpg' in images:
                rows[5].value = dir_images + '/default.jpg'
                images.remove('default.jpg')
                i = 5
            else:
                i = 4

            for image in images:
                if 'png' in image or 'jpg' in image:
                    i += 1
                    if counter <= 5:
                        counter += 1
                        rows[i].value = dir_images + '/' + image

    filename = verification(os.path.basename(file))
    wb.save(filename)
    open_file(filename)
