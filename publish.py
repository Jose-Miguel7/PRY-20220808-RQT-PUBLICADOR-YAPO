import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

import sys
import json
from openpyxl import load_workbook


def publish(excel, username, password):
    opts = Options()
    opts.add_argument("user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/71.0.3578.80 Chrome/71.0.3578.80 Safari/537.36")

    if sys.platform == "win32":
        driver = webdriver.Chrome('./chromedriver.exe', options=opts)
    else:
        driver = webdriver.Chrome('./chromedriver', options=opts)

    driver.get('https://www2.yapo.cl/login')

    email_input = driver.find_element_by_id('email_input')
    email_input.send_keys(username)

    password_input = driver.find_element_by_id('password_input')
    password_input.send_keys(password)

    submit_login = driver.find_element_by_id('submit_button')
    submit_login.click()

    wb = load_workbook(excel)
    ws = wb.active

    counter = 0
    success = 0

    for rows in ws.iter_rows(min_row=2, min_col=2):
        counter += 1

        try:
            category = str(rows[0].value)
            title = rows[1].value
            description = rows[2].value
            price_detail = str(rows[3].value)
            image1 = rows[4].value
            image2 = rows[5].value
            image3 = rows[6].value
            image4 = rows[7].value
            image5 = rows[8].value
            image6 = rows[9].value
            region_detail = str(rows[10].value)
            comuna_detail = str(rows[11].value)
            images = (image2, image3, image4, image5, image6)
        except:
            category, title, description, price_detail, image1, image2, image3 = None, None, None, None, None, None, None
            image4, image5, image6, region_detail, comuna_detail = None, None, None, None, None
        if image1:
            text_images = image1
            for image in images:
                if image:
                    text_images = text_images + ' \n ' + image

        else:
            text_images = ''

        if category and title and description and price_detail and region_detail and comuna_detail:
            driver.get('https://new.yapo.cl/legacy/ad-insert.html')
            time.sleep(2)

            ############################  SELECT CATEGORY  ############################
            with open('./data/category.json', encoding="utf8") as r:
                data = json.loads(r.read())
                data_category = data[category]
                id_2 = data_category['id_2']
                id_condition = data_category['id_condition']
                id_gender = data_category['id_gender']
                id_clothing_size = data_category['id_clothing_size']

            category_group = Select(driver.find_element_by_id('category_group'))
            category_group.select_by_value(id_2)

            if id_condition != "0" or id_gender != "0" or id_clothing_size != "0":
                if id_condition != "0":
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'condition')))
                    id_condition_select = Select(driver.find_element_by_id('condition'))
                    id_condition_select.select_by_value(id_condition)
                if id_gender != "0":
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'gender')))
                    id_gender_select = Select(driver.find_element_by_id('gender'))
                    id_gender_select.select_by_value(id_gender)
                if id_clothing_size != "0":
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'clothing_size')))
                    id_clothing_size_select = Select(driver.find_element_by_id('clothing_size'))
                    id_clothing_size_select.select_by_value(id_clothing_size)

            ############################  TITLE - DESCRIPTION - PRICE  ############################
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, 'subject')))
            subject = driver.find_element_by_id('subject')
            subject.send_keys(title)

            try:
                body = driver.find_element_by_id('body')
                body.send_keys(description)
            except:
                WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, 'body')))
                body = driver.find_element_by_id('body')
                body.send_keys(description)

            price = driver.find_element_by_id('price')
            price.send_keys(price_detail)

            ############################  REGION - COMUNA  ############################
            region = Select(driver.find_element_by_id('region'))
            region.select_by_value(region_detail)

            communes = Select(driver.find_element_by_id('communes'))
            communes.select_by_value(comuna_detail)

            ############################  IMAGES  ############################
            if text_images:
                images = driver.find_element_by_xpath("//input[@type='file']")
                images.send_keys(text_images)

            ############################  TERMS  ############################
            accept_conditions = driver.find_elements_by_class_name('iCheck-helper')[0]
            accept_conditions.click()

            ############################  UPLOAD  ############################
            time.sleep(5)
            submit_create_now = driver.find_element_by_id('submit_create_now')
            try:
                submit_create_now.click()
            except Exception as e:
                print(e)

            try:
                WebDriverWait(driver, 3).until(EC.alert_is_present(),
                                               'Timed out waiting for PA creation ' +
                                               'confirmation popup to appear.')

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except TimeoutException:
                print("no alert")

            try:
                WebDriverWait(driver, 7).until(EC.presence_of_element_located((By.ID, 'submit_create_t_prv')))
                a = driver.find_element_by_id('submit_create_t_prv')
                a.click()
            except TimeoutException:
                pass

            time.sleep(2)
            success += 1

    driver.quit()
    return counter, success
